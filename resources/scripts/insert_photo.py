#!/usr/bin/env python3
"""
insert_photo.py — Inserts a JPG/PNG image into the PHOTO area (G8:M35)
of the "Spec Sheet" worksheet in an Elite Flower recipe Excel file.

Usage:
    python3 insert_photo.py <excel_path> <image_path>

Exits 0 on success, 1 on error (error message on stderr).
Prints "OK" to stdout on success.
"""

import sys
import os
import tempfile

try:
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
    from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.units import pixels_to_EMU
    from PIL import Image as PILImage
except ImportError as e:
    print(f"ERROR: Missing Python dependency — {e}\n"
          f"Install with: pip3 install openpyxl pillow", file=sys.stderr)
    sys.exit(1)


def col_px(ws, col_letter: str) -> float:
    """Convert column width units to pixels."""
    w = ws.column_dimensions[col_letter].width
    return (w if w else 8.43) * 7.5


def row_px(ws, row_idx: int) -> float:
    """Convert row height points to pixels."""
    h = ws.row_dimensions[row_idx].height
    return (h if h else 15) * 1.3333


def is_locked_by_excel(excel_path: str) -> bool:
    """Excel writes a hidden ~$name.xlsx owner file while the workbook is open."""
    lock = os.path.join(os.path.dirname(excel_path), "~$" + os.path.basename(excel_path))
    return os.path.exists(lock)


def anchor_intersects_area(img, area_x0, area_y0, area_x1, area_y1, ws) -> bool:
    """True when an existing image overlaps the G8:M35 photo area.
    Logos and other images elsewhere on the sheet must survive."""
    try:
        anchor = img.anchor
        # AbsoluteAnchor: pos/ext in EMU (1 px = 9525 EMU)
        if hasattr(anchor, 'pos') and anchor.pos is not None:
            x = anchor.pos.x / 9525.0
            y = anchor.pos.y / 9525.0
            w = anchor.ext.cx / 9525.0 if anchor.ext else 0
            h = anchor.ext.cy / 9525.0 if anchor.ext else 0
            return x < area_x1 and (x + w) > area_x0 and y < area_y1 and (y + h) > area_y0
        # OneCell/TwoCell anchor: check the from-cell (0-based col/row)
        frm = getattr(anchor, '_from', None) or getattr(anchor, 'from_', None)
        if frm is not None:
            # G..M = cols 6..12, rows 8..35 = 7..34 (0-based)
            return 6 <= frm.col <= 12 and 7 <= frm.row <= 34
    except Exception:
        pass
    # Unknown anchor type — assume it's ours (photo area) to avoid duplicates
    return True


def insert_photo(excel_path: str, image_path: str) -> None:
    if not os.path.isfile(excel_path):
        print(f"ERROR: Excel file not found: {excel_path}", file=sys.stderr)
        sys.exit(1)

    if not os.path.isfile(image_path):
        print(f"ERROR: Image file not found: {image_path}", file=sys.stderr)
        sys.exit(1)

    if is_locked_by_excel(excel_path):
        print("ERROR: Excel file is open in Excel — close it and try again.", file=sys.stderr)
        sys.exit(1)

    try:
        wb = load_workbook(excel_path)
    except Exception as e:
        # Common cause: file locked by Excel.app / Microsoft Excel
        err = str(e)
        if 'lock' in err.lower() or 'permission' in err.lower() or 'access' in err.lower():
            print(f"ERROR: Excel file is locked — close it in Excel and try again. ({e})",
                  file=sys.stderr)
        else:
            print(f"ERROR: Cannot open Excel file: {e}", file=sys.stderr)
        sys.exit(1)

    if 'Spec Sheet' not in wb.sheetnames:
        print(f"ERROR: 'Spec Sheet' not found in {os.path.basename(excel_path)}. "
              f"Available sheets: {', '.join(wb.sheetnames)}", file=sys.stderr)
        sys.exit(1)

    ws = wb['Spec Sheet']

    # ── Calculate absolute pixel position of cell G8 ─────────────────────────
    # Sum of column widths A–F (columns 1–6)
    x_offset = sum(col_px(ws, get_column_letter(c)) for c in range(1, 7))
    # Sum of row heights 1–7
    y_offset = sum(row_px(ws, r) for r in range(1, 8))

    # ── Calculate pixel size of area G8:M35 ──────────────────────────────────
    # Columns G(7) through M(13)
    area_w = sum(col_px(ws, get_column_letter(c)) for c in range(7, 14))
    # Rows 8 through 35
    area_h = sum(row_px(ws, r) for r in range(8, 36))

    # ── Load image and compute fit while preserving aspect ratio ─────────────
    try:
        pil_img = PILImage.open(image_path).convert('RGB')
    except Exception as e:
        print(f"ERROR: Cannot open image: {e}", file=sys.stderr)
        sys.exit(1)

    orig_w, orig_h = pil_img.size
    ratio = min(area_w / orig_w, area_h / orig_h)
    new_w = int(orig_w * ratio)
    new_h = int(orig_h * ratio)

    # ── Center image inside the area ─────────────────────────────────────────
    cx = x_offset + (area_w - new_w) / 2
    cy = y_offset + (area_h - new_h) / 2

    # ── Resize to fit and save to a temp PNG ─────────────────────────────────
    tmp_file = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
    tmp_path = tmp_file.name
    tmp_file.close()

    try:
        pil_img.resize((new_w, new_h), PILImage.LANCZOS).save(tmp_path, 'PNG')

        # ── Insert using AbsoluteAnchor (pixel-precise, cell-independent) ────
        xl_img = XLImage(tmp_path)
        xl_img.width  = new_w
        xl_img.height = new_h
        xl_img.anchor = AbsoluteAnchor(
            pos=XDRPoint2D(pixels_to_EMU(int(cx)), pixels_to_EMU(int(cy))),
            ext=XDRPositiveSize2D(pixels_to_EMU(new_w), pixels_to_EMU(new_h))
        )

        # Remove ONLY images that overlap the photo area (re-insert dedup).
        # Wiping ws._images entirely would also delete logos on the sheet.
        area_x1 = x_offset + area_w
        area_y1 = y_offset + area_h
        ws._images = [
            im for im in ws._images
            if not anchor_intersects_area(im, x_offset, y_offset, area_x1, area_y1, ws)
        ]
        ws.add_image(xl_img)

        # ── Atomic save: write a sibling temp file, then replace ─────────────
        # Saving directly over the original means a crash/kill mid-save leaves
        # a truncated, permanently corrupt workbook. os.replace is atomic on
        # the same filesystem, so the original is never in a half-written state.
        tmp_xlsx = excel_path + f".tmp-{os.getpid()}.xlsx"
        try:
            wb.save(tmp_xlsx)
            os.replace(tmp_xlsx, excel_path)
        finally:
            if os.path.exists(tmp_xlsx):
                try:
                    os.unlink(tmp_xlsx)
                except OSError:
                    pass
        print("OK")

    except Exception as e:
        err = str(e)
        if 'lock' in err.lower() or 'permission' in err.lower():
            print(f"ERROR: Cannot save — file may be open in Excel. Close it and try again. ({e})",
                  file=sys.stderr)
        else:
            print(f"ERROR: Failed to insert image: {e}", file=sys.stderr)
        sys.exit(1)
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("Usage: insert_photo.py <excel_path> <image_path>", file=sys.stderr)
        sys.exit(1)

    insert_photo(sys.argv[1], sys.argv[2])
